"""
split_excel.py — Dispatcher de reparto por marca (SP1)

Lee el/los Excel "sucios" del distribuidor en entrada/ (una hoja por estado+marca),
separa cada fila de pedido por marca según el prefijo del SKU, y deposita un .xlsx
canónico y limpio en fuentes/ de cada marca. Luego regenera el data.js de las marcas
cuyo generador ya entiende el formato canónico (por ahora: GOA).

Uso:
    python split_excel.py [--input entrada] [--no-build] [--only GOA,KOM]
"""

import argparse
import datetime
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent


# ── Configuración de marcas ──────────────────────────────────────────────
@dataclass(frozen=True)
class Brand:
    name: str        # nombre legible
    code: str        # prefijo del nombre de archivo canónico
    folder: str      # carpeta destino (relativa a ROOT)
    generator: str   # script generador dentro de la carpeta
    sheet: str       # nombre de la hoja del archivo canónico
    sp1_ready: bool  # ¿su generador ya lee el formato canónico?


BRANDS = {
    "GET": Brand("San José", "San-Jose", "reporte-sanjose", "generar_data.py", "San Jose", True),
    "GOA": Brand("Garden of the Andes", "GOA", "goa-inventory", "parse_excel.py", "GOA", True),
    "KOM": Brand("Kombuchacha", "KOM", "kombuchacha", "build_data.py", "KOM", True),
    "ROB": Brand("Robinson Crusoe", "ROB", "robinson-crusoe", "build_data.py", "ROB", True),
}

HEADER = [
    "Order Date", "Order", "Product Variant", "Customer", "Salesperson",
    "Company", "Qty Delivered", "Qty Invoiced", "Qty Ordered", "Unit Price", "Total",
]

# Índices de columna (0-based)
COL_DATE, COL_ORDER, COL_VARIANT, COL_CUSTOMER, COL_SP, COL_COMPANY = 0, 1, 2, 3, 4, 5

SKU_PREFIX_RE = re.compile(r"^\s*\[([A-Z]{2,4})\d+\]")

STATE_BY_COMPANY = {
    "latinfood florida": "Florida",
    "latinfood us corp.": "Nueva York",
    "latinfood us corp": "Nueva York",
}
STATE_BY_SHEET_PREFIX = {"MIA": "Florida", "NY": "Nueva York"}


# ── Helpers puros ─────────────────────────────────────────────────────────
def sku_prefix(variant):
    """'[GOA02] Pure Chamoline' -> 'GOA'. None si no matchea o no es str."""
    if not isinstance(variant, str):
        return None
    m = SKU_PREFIX_RE.match(variant)
    return m.group(1) if m else None


def is_transaction_row(row):
    """Fila de pedido real: datetime en col A y SKU [XXX##] en col C."""
    return (
        isinstance(row[COL_DATE], datetime.datetime)
        and sku_prefix(row[COL_VARIANT]) is not None
    )


def month_key(date):
    """datetime -> 'YYYY-MM'."""
    return f"{date.year:04d}-{date.month:02d}"


def state_of(row, sheet_name):
    """Estado desde Company; respaldo por prefijo de hoja; '?' si desconocido."""
    company = row[COL_COMPANY]
    if isinstance(company, str):
        key = re.sub(r"\s+", " ", company).strip().lower()
        if key in STATE_BY_COMPANY:
            return STATE_BY_COMPANY[key]
    parts = str(sheet_name).split()
    prefix = parts[0].upper() if parts else ""
    return STATE_BY_SHEET_PREFIX.get(prefix, "?")


def find_incentive_block(rows):
    """Filas del bloque de incentivos (desde el ancla 'SOLD' al final), o [].

    Ancla: alguna celda de la fila == 'SOLD' (case-insensitive).
    """
    for i, row in enumerate(rows):
        if any(isinstance(c, str) and c.strip().upper() == "SOLD" for c in row):
            return [list(r) for r in rows[i:]]
    return []


@dataclass
class Collected:
    rows: dict         # code -> {month -> [row, ...]}
    incentives: dict   # code -> {month -> [block_row, ...]}
    unmatched: list    # [row, ...]
    stats: dict        # code -> {month -> Counter(state)}


def collect(input_files):
    """Recorre los workbooks y separa filas por marca/mes, guardando además los
    SKUs no mapeados, los bloques de incentivos y los conteos por estado."""
    rows = defaultdict(lambda: defaultdict(list))
    incentives = defaultdict(dict)
    unmatched = []
    stats = defaultdict(lambda: defaultdict(Counter))

    for path in input_files:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:  # archivo corrupto / no es xlsx
            print(f"  ⚠ No se pudo leer {Path(path).name}: {e} (se omite)")
            continue

        for ws in wb.worksheets:
            sheet_rows = list(ws.iter_rows(values_only=True))
            tx_codes, tx_months, tx_states = [], [], []

            for row in sheet_rows:
                if not is_transaction_row(row):
                    continue
                prefix = sku_prefix(row[COL_VARIANT])
                mk = month_key(row[COL_DATE])
                if prefix not in BRANDS:
                    unmatched.append(list(row))
                    continue
                rows[prefix][mk].append(list(row))
                stats[prefix][mk][state_of(row, ws.title)] += 1
                tx_codes.append(prefix)
                tx_months.append(mk)
                tx_states.append(state_of(row, ws.title))

            block = find_incentive_block(sheet_rows)
            if block and tx_codes:
                dom_code = Counter(tx_codes).most_common(1)[0][0]
                dom_month = Counter(tx_months).most_common(1)[0][0]
                dom_state = Counter(tx_states).most_common(1)[0][0]
                seg = {"state": dom_state, "rows": block}
                incentives[dom_code].setdefault(dom_month, []).append(seg)
            elif block:
                print(f"  ⚠ Bloque de incentivos en '{ws.title}' sin filas de pedido "
                      f"de marca conocida; se omite")

    # Convertir defaultdicts a dicts normales para asserts predecibles
    return Collected(
        rows={c: dict(m) for c, m in rows.items()},
        incentives={c: dict(m) for c, m in incentives.items()},
        unmatched=unmatched,
        stats={c: {mk: cnt for mk, cnt in m.items()} for c, m in stats.items()},
    )


def write_canonical(path, rows, incentive_block, sheet_name):
    """Escribe el .xlsx canónico: encabezado + filas de pedido (passthrough fiel)
    + (separador en blanco + bloque de incentivos, si lo hay)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(HEADER)
    for row in rows:
        ws.append(list(row))
    if incentive_block:
        ws.append([None] * len(HEADER))
        for row in incentive_block:
            ws.append(list(row))
    wb.save(path)


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Reparte el Excel del distribuidor por marca.")
    p.add_argument("--input", default="entrada", help="Carpeta con los .xlsx sucios.")
    p.add_argument("--no-build", action="store_true", help="No regenerar ningún data.js.")
    p.add_argument("--only", default="", help="Limitar a códigos de marca (ej. GOA,KOM).")
    return p.parse_args(argv)


def iter_input_files(input_dir):
    """Todos los .xlsx de la carpeta (sin temporales ~$), en orden de nombre."""
    d = Path(input_dir)
    if not d.is_absolute():
        d = ROOT / d
    if not d.exists():
        return []
    return sorted(p for p in d.glob("*.xlsx") if not p.name.startswith("~$"))


def run_generator(brand):
    """Corre el generador de la marca. Devuelve (returncode|None, log)."""
    folder = ROOT / brand.folder
    script = folder / brand.generator
    if not script.exists():
        return None, f"generador ausente ({brand.generator})"
    # Forzar UTF-8 en el hijo (PYTHONUTF8) y al decodificar el pipe: los
    # generadores imprimen ✓/⚠/acentos que en Windows-cp1252 saldrían corruptos
    # o harían fallar al hijo.
    env = {**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"}
    proc = subprocess.run(
        [sys.executable, brand.generator],
        cwd=folder, capture_output=True, text=True,
        encoding="utf-8", errors="replace", env=env,
    )
    return proc.returncode, proc.stdout + proc.stderr


def main(argv=None):
    # El resumen imprime ✓/⚠/↻/é; en un pipe Windows-cp1252 esto reventaría.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

    args = parse_args(argv)
    only = {c.strip().upper() for c in args.only.split(",") if c.strip()}

    files = iter_input_files(args.input)
    if not files:
        print(f"✗ No hay .xlsx en {args.input}/. Pon ahí el Excel del distribuidor.")
        return 1

    data = collect(files)
    print(f"✓ split_excel.py — {', '.join(f.name for f in files)}\n")

    build_failed = False
    for code, brand in BRANDS.items():
        if only and code not in only:
            continue
        months = data.rows.get(code, {})
        if not months:
            continue

        print(f"  {code}  ({brand.name})")
        for mk in sorted(months):
            month_rows = months[mk]
            segs = data.incentives.get(code, {}).get(mk, [])
            block = []
            for seg in segs:
                block.append(["__INCENTIVOS__", seg["state"]] + [None] * (len(HEADER) - 2))
                block.extend(seg["rows"])
                block.append([None] * len(HEADER))   # separador entre segmentos
            out = ROOT / brand.folder / "fuentes" / f"{brand.code}-{mk}.xlsx"
            write_canonical(out, month_rows, block, brand.sheet)
            st = data.stats.get(code, {}).get(mk, {})
            st_txt = " ".join(f"{k}:{v}" for k, v in st.items())
            inc_txt = " · +incentivos" if segs else ""
            rel = out.relative_to(ROOT).as_posix()
            print(f"    → {rel}   ({len(month_rows)} líneas · {st_txt}{inc_txt})")

        if not brand.sp1_ready:
            print("    ⏸ data.js NO regenerado (pendiente su sub-proyecto)")
        elif args.no_build:
            print("    ⏸ build omitido (--no-build)")
        else:
            rc, log = run_generator(brand)
            if rc == 0:
                print("    ↻ data.js regenerado ✓")
            else:
                build_failed = True
                print(f"    ✗ generador falló (rc={rc}):\n{log}")

    if data.unmatched:
        # Namespaced por mes (como los archivos de marca) para que un mes no pise
        # los SKUs sin mapear del anterior antes de que alguien los triage.
        by_month = defaultdict(list)
        for r in data.unmatched:
            by_month[month_key(r[COL_DATE])].append(r)
        written = []
        for mk in sorted(by_month):
            out = ROOT / "unmatched" / f"unmatched-{mk}.xlsx"
            write_canonical(out, by_month[mk], [], "unmatched")
            written.append(out.name)
        prefixes = sorted({sku_prefix(r[COL_VARIANT]) for r in data.unmatched})
        print(f"\n  ⚠ {len(data.unmatched)} líneas sin marca → unmatched/ "
              f"({', '.join(written)})  (prefijos: {', '.join(prefixes)})")
    else:
        print("\n  ⚠ SKUs no mapeados: (ninguno)")

    return 1 if build_failed else 0


if __name__ == "__main__":
    sys.exit(main())
