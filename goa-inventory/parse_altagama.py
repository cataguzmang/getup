"""
parse_altagama.py — Extractor del reporte de sell-through de Alta Gama (GOA)

Alta Gama entrega un reporte MUCHO más limitado que Latin Foods: una tabla con
una fila por producto y una columna por periodo, con las UNIDADES VENDIDAS de
cada periodo. No trae precios, clientes, órdenes, inventario ni fechas de
transacción, así que el análisis se limita al sell-through EN UNIDADES.

Estructura del Excel (`GOA_CM Sell Through Data_2026_YTD.xlsx`, hoja Sheet1):
  - Fila 1: `Products` | un encabezado por periodo. Los meses completos vienen
    como fecha (Excel interpreta "Jan 26" → 2026-01-26: el 26 es el AÑO), y el
    periodo parcial como texto libre ("Jul 1 - 25, 26").
  - Una fila por producto, con el formato `"<SKU> (<Nombre>)"` — el mismo SKU y
    el mismo nombre que traen los snapshots de inventario de Alta Gama, así que
    la homologación es exacta y no necesita heurísticas.
  - Última fila `TOTAL`: NO se importa como producto; se usa solo para validar
    los totales recalculados desde las filas de producto.

Este script escribe `data-altagama.js` (objeto `altaGamaData`). Es
independiente de `parse_excel.py` / `data.js` (Latin Foods): ni lee ni escribe
nada de esa fuente.

Uso:
    python parse_altagama.py
"""

import re
import json
import calendar
import datetime
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────────────
HERE = Path(__file__).parent
SOURCE_DIR = HERE / "Inventarios-GOA-altagama"
SOURCE_GLOB = "*Sell Through*.xlsx"
OUTPUT_FILE = HERE / "data-altagama.js"

BRAND = "Garden of the Andes"
DISTRIBUTOR = "Alta Gama"
METRIC = "Unidades vendidas (sell-through en unidades)"

TOTAL_LABEL = "TOTAL"          # fila de control, no es un producto
PRODUCTS_HEADER = "Products"

MESES_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]
MESES_ES_ABBR = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
                 "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

MONTHS_EN = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# "Jul 1 - 25, 26" / "Jul 1-25, 2026"  → periodo parcial
PARTIAL_RE = re.compile(
    r"^\s*([A-Za-z]{3,})\.?\s*(\d{1,2})\s*[-–—]\s*(\d{1,2})\s*,?\s*(\d{2,4})\s*$")
# "Jan 26" / "Jan-26" / "Jan 2026"     → mes completo escrito como texto
FULL_TEXT_RE = re.compile(r"^\s*([A-Za-z]{3,})\.?\s*[-\s]\s*(\d{2,4})\s*$")
# "GA Green Tea 1.4oz (Garden of the Andes - Green Tea 1.4oz)"
PRODUCT_RE = re.compile(r"^\s*(.+?)\s*\(\s*(.+?)\s*\)\s*$")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de limpieza
# ─────────────────────────────────────────────────────────────────────────────
def clean_str(value):
    """Normaliza un string: trim y colapsa espacios internos. None -> ''."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def units(value):
    """Celda de unidades → (int, None), o (None, motivo) si NO está disponible.

    Semántica confirmada con el negocio (2026-08-18): Alta Gama escribe `0`
    cuando **no hubo movimiento** en el periodo, así que un 0 es un dato real y
    se conserva como 0. Una celda vacía sí es dato ausente: se devuelve `None`
    y nunca se sustituye por cero."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, "vacía"
    if isinstance(value, bool):
        return None, "booleana"
    if isinstance(value, (int, float)):
        if value != value:                      # NaN
            return None, "NaN"
        return int(round(value)), None
    text = clean_str(value).replace(",", "")
    try:
        return int(round(float(text))), None
    except ValueError:
        return None, "no numérica (%r)" % clean_str(value)


def four_digit_year(raw):
    """26 -> 2026 · 2026 -> 2026."""
    year = int(raw)
    return 2000 + year if year < 100 else year


def full_month_period(year, month):
    last = calendar.monthrange(year, month)[1]
    return {
        "key": "%d-%02d" % (year, month),
        "label": "%s %d" % (MESES_ES[month].capitalize(), year),
        "short": MESES_ES_ABBR[month],
        "year": year,
        "month": month,
        "start": "%d-%02d-01" % (year, month),
        "end": "%d-%02d-%02d" % (year, month, last),
        "partial": False,
        "coverageDays": last,
        "monthDays": last,
    }


def partial_period(year, month, day_from, day_to):
    last = calendar.monthrange(year, month)[1]
    return {
        "key": "%d-%02d" % (year, month),
        "label": "%s %d" % (MESES_ES[month].capitalize(), year),
        "short": "%s 1–%d" % (MESES_ES_ABBR[month], day_to),
        "year": year,
        "month": month,
        "start": "%d-%02d-%02d" % (year, month, day_from),
        "end": "%d-%02d-%02d" % (year, month, day_to),
        "partial": True,
        "coverageDays": day_to - day_from + 1,
        "monthDays": last,
    }


def parse_period_header(value):
    """Encabezado de columna → periodo. Lanza ValueError si no se reconoce.

    Excel guarda "Jan 26" como fecha 2026-01-26 (el 26 es el AÑO, no el día):
    de ahí se toman año y mes, y el periodo se expande al mes completo.
    """
    if isinstance(value, (datetime.datetime, datetime.date)):
        return full_month_period(value.year, value.month)

    text = clean_str(value)
    if not text:
        raise ValueError("encabezado de periodo vacío")

    m = PARTIAL_RE.match(text)
    if m:
        name, day_from, day_to, year = m.groups()
        month = MONTHS_EN.get(name[:3].lower())
        if month is None:
            raise ValueError("mes no reconocido en %r" % text)
        return partial_period(four_digit_year(year), month, int(day_from), int(day_to))

    m = FULL_TEXT_RE.match(text)
    if m:
        name, year = m.groups()
        month = MONTHS_EN.get(name[:3].lower())
        if month is None:
            raise ValueError("mes no reconocido en %r" % text)
        return full_month_period(four_digit_year(year), month)

    raise ValueError("encabezado de periodo no reconocido: %r" % text)


def split_product(value):
    """'GA Green Tea 1.4oz (Garden of the Andes - Green Tea 1.4oz)'
       -> ('GA Green Tea 1.4oz', 'Garden of the Andes - Green Tea 1.4oz').

    Si el texto no trae el paréntesis con el nombre largo, el SKU queda en None
    y el texto completo se usa como nombre: no se inventa un SKU."""
    text = clean_str(value)
    m = PRODUCT_RE.match(text)
    if not m:
        return None, text
    return clean_str(m.group(1)), clean_str(m.group(2))


def short_name(name, sku):
    """Etiqueta corta para gráficos: 'Garden of the Andes - Green Tea 1.4oz'
       -> 'Green Tea 1.4oz'."""
    base = name.split(" - ", 1)[1] if " - " in name else name
    base = clean_str(base)
    return base or clean_str(sku) or name


def pct(part, whole, digits=1):
    return round(part / whole * 100, digits) if whole else None


# ─────────────────────────────────────────────────────────────────────────────
# Lectura
# ─────────────────────────────────────────────────────────────────────────────
def find_source(folder=SOURCE_DIR, pattern=SOURCE_GLOB):
    """Excel de sell-through más reciente por nombre; ignora temporales (~$)."""
    folder = Path(folder)
    if not folder.exists():
        return None
    matches = sorted(p for p in folder.glob(pattern) if not p.name.startswith("~$"))
    return matches[-1] if matches else None


def read_sheet_rows(path):
    """Filas (values_only) de la hoja de datos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    return list(ws.iter_rows(values_only=True))


def parse_sheet(rows):
    """Filas crudas → (periodos, productos, fila TOTAL, incidencias).

    La fila `TOTAL` se devuelve aparte: nunca entra como producto.
    """
    if not rows:
        raise RuntimeError("La hoja está vacía.")

    header = rows[0]
    if clean_str(header[0]).lower() != PRODUCTS_HEADER.lower():
        raise RuntimeError(
            "Se esperaba %r en A1 y se encontró %r."
            % (PRODUCTS_HEADER, clean_str(header[0])))

    periods, columns = [], []
    for idx, cell in enumerate(header[1:], start=1):
        if cell is None or clean_str(cell) == "":
            continue
        periods.append(parse_period_header(cell))
        columns.append(idx)

    if not periods:
        raise RuntimeError("No se reconoció ninguna columna de periodo.")

    keys = [p["key"] for p in periods]
    if len(set(keys)) != len(keys):
        raise RuntimeError("Periodos duplicados en el encabezado: %s" % keys)

    products, total_row, issues = [], None, []
    for row_no, row in enumerate(rows[1:], start=2):
        raw = clean_str(row[0]) if row else ""
        if not raw:
            continue

        values = {}
        for key, col in zip(keys, columns):
            value, reason = units(row[col] if col < len(row) else None)
            values[key] = value
            if reason:
                issues.append("fila %d (%s) · %s: celda %s" % (row_no, raw, key, reason))

        if raw.upper() == TOTAL_LABEL:
            total_row = values
            continue

        sku, name = split_product(raw)
        products.append({
            "sku": sku,
            "name": name,
            "shortName": short_name(name, sku),
            "byPeriod": values,
        })

    if not products:
        raise RuntimeError("No se encontró ninguna fila de producto.")

    skus = [p["sku"] for p in products if p["sku"]]
    dup = sorted({s for s in skus if skus.count(s) > 1})
    if dup:
        raise RuntimeError("SKU duplicados en el Excel: %s" % dup)

    names = [p["name"].lower() for p in products]
    dup_names = sorted({n for n in names if names.count(n) > 1})
    if dup_names:
        raise RuntimeError("Productos duplicados por nombre: %s" % dup_names)

    return periods, products, total_row, issues


# ─────────────────────────────────────────────────────────────────────────────
# Agregación
# ─────────────────────────────────────────────────────────────────────────────
def sum_units(values):
    """Suma ignorando los None (dato no disponible ≠ cero)."""
    return sum(v for v in values if v is not None)


def build_report(path):
    rows = read_sheet_rows(path)
    periods, products, total_row, issues = parse_sheet(rows)

    keys = [p["key"] for p in periods]
    full_keys = [p["key"] for p in periods if not p["partial"]]
    partial_keys = [p["key"] for p in periods if p["partial"]]

    # Totales por periodo, recalculados SIEMPRE desde las filas de producto
    for period in periods:
        period["units"] = sum_units(p["byPeriod"][period["key"]] for p in products)

    # Validación contra la fila TOTAL del Excel
    checks, all_match = [], True
    for period in periods:
        expected = (total_row or {}).get(period["key"])
        ok = expected is not None and expected == period["units"]
        all_match = all_match and ok
        checks.append({"key": period["key"], "label": period["label"],
                       "excelTotal": expected, "computed": period["units"], "ok": ok})

    grand_total = sum(p["units"] for p in periods)
    total_full = sum(p["units"] for p in periods if not p["partial"])

    # Variación mes a mes: solo entre meses COMPLETOS consecutivos. Un periodo
    # parcial no se compara con un mes completo (no son magnitudes homologables).
    prev = None
    for period in periods:
        if period["partial"]:
            period["delta"] = None
            period["deltaPct"] = None
            continue
        period["delta"] = None if prev is None else period["units"] - prev
        period["deltaPct"] = (None if prev in (None, 0)
                              else pct(period["units"] - prev, prev))
        prev = period["units"]

    # Productos
    for prod in products:
        prod["units"] = sum_units(prod["byPeriod"].values())
        prod["unitsFullMonths"] = sum_units(prod["byPeriod"][k] for k in full_keys)
        prod["share"] = pct(prod["units"], grand_total)
        prod["avgFullMonth"] = (round(prod["unitsFullMonths"] / len(full_keys), 1)
                                if full_keys else None)
        prod["missingPeriods"] = [k for k in keys if prod["byPeriod"][k] is None]
        best_key = max((k for k in full_keys if prod["byPeriod"][k] is not None),
                       key=lambda k: prod["byPeriod"][k], default=None)
        best_period = next((p for p in periods if p["key"] == best_key), None)
        prod["best"] = None if best_period is None else {
            "key": best_key, "label": best_period["label"],
            "units": prod["byPeriod"][best_key]}

    products.sort(key=lambda p: (-p["units"], p["shortName"]))

    top_product = products[0] if products else None
    full_periods = [p for p in periods if not p["partial"]]
    best_full = max(full_periods, key=lambda p: p["units"], default=None)
    partial = next((p for p in periods if p["partial"]), None)

    period_start = min(p["start"] for p in periods)
    period_end = max(p["end"] for p in periods)
    period_label = ("%s – %s" % (periods[0]["label"], periods[-1]["label"])
                    if len(periods) > 1 else periods[0]["label"])
    if partial:
        period_label += " (al %d)" % int(partial["end"][-2:])

    if partial:
        july_note = ("Los datos de %s corresponden al periodo comprendido entre "
                     "el %d y el %d de %s de %d."
                     % (partial["label"].split(" ")[0].lower(),
                        int(partial["start"][-2:]), int(partial["end"][-2:]),
                        MESES_ES[partial["month"]], partial["year"]))
    else:
        july_note = "Todos los periodos del reporte son meses completos."

    return {
        "meta": {
            "brand": BRAND,
            "distributor": DISTRIBUTOR,
            "metric": METRIC,
            "source": path.name,
            "periodStart": period_start,
            "periodEnd": period_end,
            "periodLabel": period_label,
            "cutoffDate": period_end,
            "partialPeriods": partial_keys,
            "notes": [
                "Alta Gama proporciona información de sell-through en unidades. "
                "El reporte no incluye precios ni valores monetarios.",
                july_note,
            ],
            "validation": {
                "totalRowPresent": total_row is not None,
                "allMatch": all_match,
                "checks": checks,
                "issues": issues,
            },
        },
        "totals": {
            "units": grand_total,
            "unitsFullMonths": total_full,
            "products": len(products),
            "periods": len(periods),
            "fullMonths": len(full_keys),
            "partialPeriods": len(partial_keys),
            "avgFullMonth": round(total_full / len(full_keys), 1) if full_keys else None,
        },
        "highlights": {
            "topProduct": None if not top_product else {
                "sku": top_product["sku"], "shortName": top_product["shortName"],
                "units": top_product["units"], "share": top_product["share"]},
            "bestFullMonth": None if not best_full else {
                "key": best_full["key"], "label": best_full["label"],
                "units": best_full["units"]},
            "partial": None if not partial else {
                "key": partial["key"], "label": partial["label"],
                "short": partial["short"], "units": partial["units"],
                "start": partial["start"], "end": partial["end"],
                "coverageDays": partial["coverageDays"],
                "monthDays": partial["monthDays"]},
        },
        "periods": periods,
        "products": products,
    }


def main():
    path = find_source()
    if path is None:
        raise SystemExit("✗ No se encontró ningún '%s' en %s." % (SOURCE_GLOB, SOURCE_DIR))

    report = build_report(path)
    js = "const altaGamaData = " + json.dumps(report, indent=2, ensure_ascii=False) + ";\n"
    OUTPUT_FILE.write_text(js, encoding="utf-8")

    m, t, v = report["meta"], report["totals"], report["meta"]["validation"]
    print("✓ data-altagama.js generado")
    print("  Distribuidor: %s" % m["distributor"])
    print("  Fuente:       %s" % m["source"])
    print("  Periodo:      %s (%s → %s)" % (m["periodLabel"], m["periodStart"], m["periodEnd"]))
    print("  Productos:    %d   Periodos: %d (%d completos + %d parcial)"
          % (t["products"], t["periods"], t["fullMonths"], t["partialPeriods"]))
    print("  Unidades:     %d total · %d en meses completos"
          % (t["units"], t["unitsFullMonths"]))
    for c in v["checks"]:
        print("    %s %s: calculado %s · Excel TOTAL %s"
              % ("✓" if c["ok"] else "✗", c["key"], c["computed"], c["excelTotal"]))
    if not v["allMatch"]:
        print("  ⚠ Los totales calculados NO cuadran con la fila TOTAL del Excel.")
    if v["issues"]:
        print("  ⚠ Celdas no utilizables (tratadas como dato no disponible):")
        for i in v["issues"]:
            print("    · %s" % i)


if __name__ == "__main__":
    main()
