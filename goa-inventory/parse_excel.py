"""
parse_excel.py — Extractor de datos del reporte GOA (Garden of the Andes · LatinFood US Corp)

Lee el formato DEFINITIVO del distribuidor (Latin-GOA-MAY.xlsx), un export
transaccional donde cada fila es una línea de pedido, lo limpia y agrega, y
exporta `data.js` (objeto `reportData`) listo para alimentar el dashboard.

Estructura del Excel nuevo (hoja "GOA"):
  - Encabezado (fila 1): Order Date | Order | Product Variant | Customer |
    Salesperson | Company | Qty Delivered | Qty Invoiced | Qty Ordered |
    Unit Price | Total
  - Las líneas de pedido están agrupadas por vendedor con filas-subtotal
    intercaladas (ej. "Mireya Fernandez (40)"). Esas filas se IGNORAN: los
    totales se recalculan desde el detalle para evitar inconsistencias.
  - Al final hay un bloque manual de incentivos/muestras (SOLD/FREE/...),
    que se parsea de forma tolerante.

Uso:
    python parse_excel.py
"""

import re
import json
import datetime
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────────────
HERE = Path(__file__).parent
SOURCES_DIR = HERE / "fuentes"
OUTPUT_FILE = HERE / "data.js"
SHEET_CANDIDATES = ("GOA", "Sheet1")  # orden de preferencia; si no, la primera hoja

BRAND = "Garden of the Andes"
DISTRIBUTOR = "LatinFood US Corp"

# El SKU viene embebido en el nombre del producto: "[GOA01] Green Tea Herbal Tea"
SKU_PATTERN = re.compile(r"^\s*\[(GOA\d+)\]\s*(.*)$")

# Índices de columnas (0-based) en el bloque transaccional
COL_DATE, COL_ORDER, COL_VARIANT, COL_CUSTOMER, COL_SALESPERSON = 0, 1, 2, 3, 4
COL_COMPANY, COL_QTY_DELIVERED, COL_QTY_INVOICED, COL_QTY_ORDERED = 5, 6, 7, 8
COL_UNIT_PRICE, COL_TOTAL = 9, 10

MESES_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de limpieza
# ─────────────────────────────────────────────────────────────────────────────
def num(value):
    """Convierte un valor a número (float) de forma segura. None/texto/NaN -> 0."""
    if value is None:
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return 0 if value != value else value  # descarta NaN
    return 0


def clean_str(value):
    """Normaliza un string: trim y colapsa espacios internos. None -> ''."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def split_sku(variant):
    """'[GOA01] Green Tea Herbal Tea' -> ('GOA01', 'Green Tea Herbal Tea')."""
    m = SKU_PATTERN.match(str(variant))
    if not m:
        return None, clean_str(variant)
    return m.group(1), clean_str(m.group(2))


def month_key(date_iso):
    """'2026-02-24' -> '2026-02'."""
    return date_iso[:7]


def month_label(key):
    """'2026-02' -> 'Febrero 2026'."""
    y, m = key.split("-")
    return f"{MESES_ES[int(m)].capitalize()} {y}"


def list_source_files(folder):
    """Todos los .xlsx de la carpeta, en orden de nombre, sin temporales (~$)."""
    folder = Path(folder)
    if not folder.exists():
        return []
    return sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~$"))


def read_sheet_rows(path):
    """Filas (values_only) de la hoja preferida: GOA → Sheet1 → primera."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    return list(ws.iter_rows(values_only=True))


def empty_incentives():
    return {"freeUnitsBySalesperson": [], "costItems": [], "totalDescuentos": 0.0}


def merge_incentives(incs):
    """Combina varios bloques de incentivos en uno."""
    free, cost, total = {}, [], 0.0
    for inc in incs:
        for f in inc.get("freeUnitsBySalesperson", []):
            free[f["name"]] = free.get(f["name"], 0) + f["free"]
        cost.extend(inc.get("costItems", []))
        total += inc.get("totalDescuentos", 0.0)
    return {
        "freeUnitsBySalesperson": [{"name": n, "free": v} for n, v in free.items()],
        "costItems": cost,
        "totalDescuentos": round(total, 2),
    }


def dominant_month(lines):
    """Mes más frecuente entre las líneas (para atribuir incentivos del archivo)."""
    counts = {}
    for ln in lines:
        counts[ln["month"]] = counts.get(ln["month"], 0) + 1
    return max(counts, key=counts.get)


def load_all_sources(folder):
    """Lee todas las fuentes, dedup por orden entre archivos, e incentivos por mes.

    Devuelve: (lines, incentives_by_month, source_names, dup_orders).
    La PRIMERA aparición de un número de orden gana; las demás se reportan.
    """
    seen_orders = set()
    lines = []
    incentives_by_month = {}
    dup_orders = []
    sources = []

    for path in list_source_files(folder):
        sources.append(path.name)
        try:
            rows = read_sheet_rows(path)
        except Exception as e:                       # archivo corrupto / sin hoja
            print(f"  ⚠ No se pudo leer {path.name}: {e} (se omite)")
            continue
        file_lines = parse_transactions(rows)
        kept = []
        for ln in file_lines:
            if ln["order"] in seen_orders:
                dup_orders.append(ln["order"])
                continue
            kept.append(ln)
        for ln in kept:
            seen_orders.add(ln["order"])
        lines.extend(kept)

        inc = parse_incentives(rows)
        has_inc = inc["freeUnitsBySalesperson"] or inc["costItems"] or inc["totalDescuentos"]
        if has_inc and kept:
            m = dominant_month(kept)
            incentives_by_month[m] = merge_incentives(
                [incentives_by_month.get(m, empty_incentives()), inc])

    return lines, incentives_by_month, sources, sorted(set(dup_orders))


def is_transaction_row(row):
    """
    Una línea de pedido real tiene fecha en col A y un SKU [GOAxx] en la col
    'Product Variant'. Esto descarta filas-subtotal de vendedor, blancos y el
    bloque de incentivos del final.
    """
    return (
        isinstance(row[COL_DATE], datetime.datetime)
        and isinstance(row[COL_VARIANT], str)
        and SKU_PATTERN.match(row[COL_VARIANT])
    )


# ─────────────────────────────────────────────────────────────────────────────
# Parsing del bloque transaccional
# ─────────────────────────────────────────────────────────────────────────────
def parse_transactions(rows):
    """Devuelve la lista de líneas de pedido limpias (dicts)."""
    lines = []
    for row in rows:
        if not is_transaction_row(row):
            continue
        sku, name = split_sku(row[COL_VARIANT])
        if sku is None:
            continue
        date = row[COL_DATE]
        lines.append({
            "date": date.date().isoformat(),
            "month": month_key(date.date().isoformat()),
            "order": clean_str(row[COL_ORDER]),
            "sku": sku,
            "product": name,
            "customer": clean_str(row[COL_CUSTOMER]),
            "salesperson": clean_str(row[COL_SALESPERSON]),
            "qtyOrdered": int(num(row[COL_QTY_ORDERED])),
            "qtyDelivered": int(num(row[COL_QTY_DELIVERED])),
            "qtyInvoiced": int(num(row[COL_QTY_INVOICED])),
            "unitPrice": round(num(row[COL_UNIT_PRICE]), 2),
            "total": round(num(row[COL_TOTAL]), 2),
        })
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Agregaciones
# ─────────────────────────────────────────────────────────────────────────────
def aggregate_products(lines):
    """Agrega por SKU, con desglose por vendedor."""
    prods = {}
    for ln in lines:
        p = prods.setdefault(ln["sku"], {
            "sku": ln["sku"], "name": ln["product"],
            "units": 0, "revenue": 0.0, "_orders": set(), "_sp": {},
        })
        p["units"] += ln["qtyOrdered"]
        p["revenue"] += ln["total"]
        p["_orders"].add(ln["order"])
        sp = p["_sp"].setdefault(ln["salesperson"], {"units": 0, "revenue": 0.0})
        sp["units"] += ln["qtyOrdered"]
        sp["revenue"] += ln["total"]

    out = []
    for p in prods.values():
        salespeople = [
            {"name": n, "units": v["units"], "revenue": round(v["revenue"], 2)}
            for n, v in p["_sp"].items()
        ]
        salespeople.sort(key=lambda s: (-s["units"], -s["revenue"]))
        out.append({
            "sku": p["sku"], "name": p["name"],
            "units": p["units"], "revenue": round(p["revenue"], 2),
            "orders": len(p["_orders"]),
            "salespeople": salespeople,
        })
    out.sort(key=lambda p: (-p["units"], -p["revenue"]))
    return out


def aggregate_salespeople(lines):
    """Agrega por vendedor: unidades, ingresos, órdenes y clientes distintos."""
    sps = {}
    for ln in lines:
        s = sps.setdefault(ln["salesperson"], {
            "name": ln["salesperson"], "units": 0, "revenue": 0.0,
            "_orders": set(), "_customers": set(),
        })
        s["units"] += ln["qtyOrdered"]
        s["revenue"] += ln["total"]
        s["_orders"].add(ln["order"])
        s["_customers"].add(ln["customer"])

    out = [{
        "name": s["name"], "units": s["units"], "revenue": round(s["revenue"], 2),
        "orders": len(s["_orders"]), "customers": len(s["_customers"]),
    } for s in sps.values()]
    out.sort(key=lambda s: (-s["revenue"], -s["units"]))
    return out


def aggregate_customers(lines):
    """Agrega por cliente."""
    cs = {}
    for ln in lines:
        c = cs.setdefault(ln["customer"], {
            "name": ln["customer"], "units": 0, "revenue": 0.0, "_orders": set(),
        })
        c["units"] += ln["qtyOrdered"]
        c["revenue"] += ln["total"]
        c["_orders"].add(ln["order"])

    out = [{
        "name": c["name"], "units": c["units"], "revenue": round(c["revenue"], 2),
        "orders": len(c["_orders"]),
    } for c in cs.values()]
    out.sort(key=lambda c: (-c["revenue"], -c["units"]))
    return out


def aggregate_daily(lines):
    """Serie temporal: unidades, ingresos y órdenes por fecha."""
    days = {}
    for ln in lines:
        d = days.setdefault(ln["date"], {
            "date": ln["date"], "units": 0, "revenue": 0.0, "_orders": set(),
        })
        d["units"] += ln["qtyOrdered"]
        d["revenue"] += ln["total"]
        d["_orders"].add(ln["order"])

    out = [{
        "date": d["date"], "units": d["units"],
        "revenue": round(d["revenue"], 2), "orders": len(d["_orders"]),
    } for d in days.values()]
    out.sort(key=lambda d: d["date"])
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Parsing del bloque de incentivos (manual / irregular → tolerante a fallos)
# ─────────────────────────────────────────────────────────────────────────────
def parse_incentives(rows):
    """
    Extrae lo que se pueda del bloque manual del final:
      - cajas gratis (FREE) por vendedor,
      - ítems de costo etiquetados (muestras, nuevos clientes, ...),
      - total de descuentos (último número suelto del bloque).

    Es defensivo a propósito: si el bloque cambia o falta, devuelve listas
    vacías en vez de romper el pipeline.
    """
    free_by_sp = []
    cost_items = []
    total_descuentos = 0.0

    # Localiza la fila-encabezado del bloque (col B == 'SOLD')
    start = None
    for i, row in enumerate(rows):
        if clean_str(row[COL_ORDER]).upper() == "SOLD":
            start = i
            break
    if start is None:
        return {"freeUnitsBySalesperson": [], "costItems": [], "totalDescuentos": 0.0}

    # En ese encabezado: B=SOLD, C=FREE → localiza la columna FREE dinámicamente
    header = rows[start]
    free_col = COL_VARIANT  # por defecto C
    for c, cell in enumerate(header):
        if clean_str(cell).upper() == "FREE":
            free_col = c
            break

    for row in rows[start + 1:]:
        label = clean_str(row[COL_DATE]) if len(row) else ""
        last_num = _last_number(row)

        # Fila de vendedor con cajas gratis: nombre en col A + número en FREE
        free_val = num(row[free_col]) if len(row) > free_col else 0
        is_sp_row = bool(label) and not _looks_like_cost_label(label)
        if is_sp_row and free_val > 0:
            free_by_sp.append({"name": label, "free": int(free_val)})

        # Ítem de costo etiquetado (texto descriptivo + monto al final)
        cost_label = _find_cost_label(row)
        if cost_label and last_num:
            cost_items.append({"label": cost_label, "amount": round(last_num, 2)})

        # El total de descuentos suele ser el último número suelto del bloque
        if last_num:
            total_descuentos = round(last_num, 2)

    return {
        "freeUnitsBySalesperson": free_by_sp,
        "costItems": cost_items,
        "totalDescuentos": total_descuentos,
    }


def _last_number(row):
    """Último valor numérico de una fila (de derecha a izquierda)."""
    for cell in reversed(list(row)):
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            if cell == cell:  # no NaN
                return float(cell)
    return None


def _looks_like_cost_label(text):
    t = text.upper()
    return any(k in t for k in ("CAJA", "MUESTRA", "NEW CUSTOMER", "DESCUENTO", "INCENTIVO"))


def _find_cost_label(row):
    """Busca en la fila un texto descriptivo de costo (muestras, cajas, etc.)."""
    for cell in row:
        s = clean_str(cell)
        if s and _looks_like_cost_label(s):
            return s
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Orquestación
# ─────────────────────────────────────────────────────────────────────────────
def build_report():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    lines = parse_transactions(rows)
    if not lines:
        raise RuntimeError("No se detectaron líneas de transacción. ¿Cambió el formato?")

    products = aggregate_products(lines)
    salespeople = aggregate_salespeople(lines)
    customers = aggregate_customers(lines)
    daily = aggregate_daily(lines)
    incentives = parse_incentives(rows)

    # Inyecta cajas gratis en cada vendedor (cruce con bloque de incentivos)
    free_map = {f["name"].lower(): f["free"] for f in incentives["freeUnitsBySalesperson"]}
    for s in salespeople:
        # match tolerante por prefijo (el bloque usa nombres abreviados)
        s["freeUnits"] = next(
            (v for k, v in free_map.items() if k in s["name"].lower() or s["name"].lower().startswith(k[:8])),
            0,
        )

    dates = sorted({ln["date"] for ln in lines})
    period_start, period_end = dates[0], dates[-1]
    start_month = int(period_start.split("-")[1])
    period_label = f"{MESES_ES[start_month].capitalize()} {period_start.split('-')[0]}"

    totals = {
        "units": sum(p["units"] for p in products),
        "revenue": round(sum(p["revenue"] for p in products), 2),
        "orders": len({ln["order"] for ln in lines}),
        "customers": len({ln["customer"] for ln in lines}),
        "salespeople": len(salespeople),
        "freeUnits": sum(f["free"] for f in incentives["freeUnitsBySalesperson"]),
    }

    return {
        "meta": {
            "brand": BRAND,
            "distributor": DISTRIBUTOR,
            "periodStart": period_start,
            "periodEnd": period_end,
            "periodLabel": period_label,
            "lineItems": len(lines),
            "source": EXCEL_FILE.name,
        },
        "totals": totals,
        "products": products,
        "salespeople": salespeople,
        "customers": customers,
        "daily": daily,
        "incentives": incentives,
    }


def main():
    report = build_report()
    js = "const reportData = " + json.dumps(report, indent=2, ensure_ascii=False) + ";\n"
    OUTPUT_FILE.write_text(js, encoding="utf-8")
    t = report["totals"]
    print("✓ data.js generado")
    print(f"  Periodo:    {report['meta']['periodLabel']} "
          f"({report['meta']['periodStart']} → {report['meta']['periodEnd']})")
    print(f"  Líneas:     {report['meta']['lineItems']}")
    print(f"  Productos:  {len(report['products'])} SKUs")
    print(f"  Vendedores: {len(report['salespeople'])}")
    print(f"  Clientes:   {t['customers']}")
    print(f"  Órdenes:    {t['orders']}")
    print(f"  Unidades:   {t['units']}")
    print(f"  Ingresos:   ${t['revenue']:,.2f}")
    print(f"  Cajas gratis (incentivos): {t['freeUnits']}")


if __name__ == "__main__":
    main()
