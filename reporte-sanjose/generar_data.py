"""
generar_data.py
Lee los Excel canónicos de fuentes/ (dejados por split_excel.py y por la
migración del histórico) y genera data.js para el reporte HTML de San José.
El estado (Florida / Nueva York) se deriva de la columna Company.

Uso: python generar_data.py
"""

import re
import json
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl

HERE = Path(__file__).parent
SOURCES_DIR = HERE / "fuentes"
OUTPUT_FILE = HERE / "data.js"

MES_CORTO = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
             7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
MES_LARGO = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',
             6:'Junio',7:'Julio',8:'Agosto',9:'Septiembre',
             10:'Octubre',11:'Noviembre',12:'Diciembre'}

# Formato canónico: Order Date · Order · Product Variant · Customer · Salesperson ·
# Company · Qty Delivered · Qty Invoiced · Qty Ordered · Unit Price · Total
SKU_RE = re.compile(r"^\s*\[[A-Z]{2,4}\d+\]\s*")
STATE_BY_COMPANY = {
    "latinfood florida": "Florida",
    "latinfood us corp.": "Nueva York",
    "latinfood us corp": "Nueva York",
}


def month_key(date):
    return (date.year, date.month)

def month_label_short(ym):
    y, m = ym
    return f"{MES_CORTO[m]} {str(y)[2:]}"

def month_label_long(ym):
    y, m = ym
    return f"{MES_LARGO[m]} {y}"

def month_label_btn(ym):
    y, m = ym
    return f"{MES_CORTO[m]} {y}"


def _source_files():
    if not SOURCES_DIR.exists():
        return []
    return sorted(p for p in SOURCES_DIR.glob("*.xlsx") if not p.name.startswith("~$"))


def _state_from_company(company):
    if not isinstance(company, str):
        return None
    return STATE_BY_COMPANY.get(re.sub(r"\s+", " ", company).strip().lower())


def _is_tx(row):
    """Fila de pedido real: datetime en col A y SKU [XXX##] en col C (Product Variant)."""
    return (
        isinstance(row[0], datetime)
        and isinstance(row[2], str)
        and SKU_RE.match(row[2]) is not None
    )


def _product_name(variant):
    """'[GET01] Jack Mackerel in Brine ' -> 'Jack Mackerel in Brine'."""
    return SKU_RE.sub("", str(variant)).strip()


def _detect_orientation(raw):
    """Dado un grupo de filas crudas del mismo mes (tuplas (qty, total, unit_price)),
    decide si viene 'normal' (Total = qty*UnitPrice) o 'swap' (UnitPrice = qty*Total).
    Usa solo filas inequívocas (qty>1 y ambos montos != 0). Ante la duda, 'normal'."""
    normal = swap = 0
    for qty, total, unit in raw:
        if qty and qty > 1 and total and unit:
            if abs(total - qty * unit) < 0.01:
                normal += 1
            elif abs(unit - qty * total) < 0.01:
                swap += 1
    return "swap" if swap > normal else "normal"


def load_rows():
    """Lee todos los fuentes/*.xlsx (canónico), detecta la orientación de montos por
    mes y devuelve las filas de pedido limpias. Ignora incentivos/separadores."""
    raw_by_month = defaultdict(list)   # (y,m) -> [row_tuple]
    unmapped = set()
    for path in _source_files():
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if not _is_tx(row):
                continue
            raw_by_month[(row[0].year, row[0].month)].append(row)

    rows = []
    for ym, month_rows in raw_by_month.items():
        orient = _detect_orientation(
            [(r[6], r[10], r[9]) for r in month_rows])  # (qty_del, total, unit_price)
        for row in month_rows:
            date, order, variant, customer, salesperson, company = row[:6]
            qty_del, unit_price, total = row[6], row[9], row[10]
            state = _state_from_company(company)
            if not state:
                label = company.strip() if isinstance(company, str) and company.strip() else "(sin Company)"
                unmapped.add(label)
                continue
            qty = qty_del or 0
            if qty <= 0:
                continue
            a = total or 0
            b = unit_price or 0
            if orient == "swap":
                line_rev = b            # Unit Price es el total de línea
                box_price = a           # Total es el precio por caja
            else:
                line_rev = a            # normal
                box_price = b
            is_promo = (line_rev == 0)  # caja gratis: total de línea 0 (entrega > 0)
            rows.append({
                'company': company or '',
                'state': state,
                'customer': (customer or '').strip(),
                'order': order or '',
                'date': date,
                'product': _product_name(variant),
                'qty': qty,
                'line_rev': line_rev,
                'box_price': box_price,
                'is_promo': is_promo,
                'salesperson': (salesperson or '').strip(),
            })
    if unmapped:
        print(f"  ⚠ Company sin estado mapeado (filas omitidas): {', '.join(sorted(unmapped))}")
    return rows


def _parse_discount_label(text):
    """'In Brine == > 3 cases' -> ('In Brine', 3). None si no reconoce producto."""
    t = str(text).lower()
    product = None
    if "in brine" in t:
        product = "In Brine"
    elif "tomate" in t or "tomato" in t:
        product = "Tomate"
    if product is None:
        return None
    mnum = re.search(r"(\d+)\s*cases", t)
    cases = int(mnum.group(1)) if mnum else 0
    return product, cases


def _last_number(row):
    for cell in reversed(list(row)):
        if isinstance(cell, (int, float)) and not isinstance(cell, bool) and cell == cell:
            return float(cell)
    return None


def load_incentivos():
    """Lee los bloques `__INCENTIVOS__` de los fuentes/ y los parsea por (mes, estado).
    Defensivo: el bloque es manual e irregular; ante lo que no reconoce, deja vacío."""
    out = {}
    for path in _source_files():
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        allrows = list(ws.iter_rows(values_only=True))
        tx_months = [(r[0].year, r[0].month) for r in allrows if _is_tx(r)]
        if not tx_months:
            continue
        # mes dominante; desempate determinista por el mes más reciente
        dom_month = max(set(tx_months), key=lambda ym: (tx_months.count(ym), ym))
        i = 0
        while i < len(allrows):
            row = allrows[i]
            if isinstance(row[0], str) and row[0].strip() == "__INCENTIVOS__":
                state = row[1]
                seg = []
                i += 1
                while i < len(allrows) and not (
                        isinstance(allrows[i][0], str) and allrows[i][0].strip() == "__INCENTIVOS__"):
                    seg.append(allrows[i])
                    i += 1
                out.setdefault(dom_month, {})[state] = _parse_incentive_segment(seg)
            else:
                i += 1
    return out


def _parse_incentive_segment(seg):
    """Parsea un segmento de incentivos en descuentos / free / sold / vendedores nuevos."""
    descuentos = []
    free_rep, sold_rep = {}, {}
    vendedores_nuevos = []
    in_new_customers = False
    for row in seg:
        col_a = row[0].strip() if isinstance(row[0], str) else row[0]
        if isinstance(col_a, str) and col_a.upper() == "NEW CUSTOMERS":
            in_new_customers = True
            continue
        if in_new_customers:
            if isinstance(col_a, str) and col_a:
                vendedores_nuevos.append(col_a)
            continue
        if isinstance(col_a, str) and col_a and col_a.upper() != "SOLD":
            sold = row[1] if isinstance(row[1], (int, float)) and not isinstance(row[1], bool) else None
            free = row[2] if isinstance(row[2], (int, float)) and not isinstance(row[2], bool) else None
            if sold is not None:
                sold_rep[col_a] = int(round(sold))
            if free is not None:
                free_rep[col_a] = int(round(free))
        # La etiqueta de descuento vive en el área "Descuentos" (col D en adelante), NO
        # en col A (nombre del vendedor) — así un nombre con "tomate/in brine" no se
        # confunde con un producto.
        matched = False
        for cell in row[3:]:
            lab = _parse_discount_label(cell) if isinstance(cell, str) else None
            if lab:
                amount = _last_number(row) or 0.0
                descuentos.append({"product": lab[0], "cases": lab[1], "amount": round(amount, 2)})
                matched = True
                break
        # Aviso si una etiqueta parece descuento ("... N cases") pero no reconocemos el
        # producto: se estaría subcontando el descuento (y sobrestimando el revenue neto).
        if not matched:
            for cell in row[3:]:
                if isinstance(cell, str) and re.search(r"\d+\s*cases", cell.lower()):
                    print(f"  ⚠ Descuento con producto no reconocido: {cell!r} (no sumado)")
                    break
    total = round(sum(d["amount"] for d in descuentos), 2)
    return {
        "descuentos": descuentos,
        "totalDescuentos": total,
        "freeReportado": free_rep,
        "soldReportado": sold_rep,
        "vendedoresNuevos": vendedores_nuevos,
    }


def new_customers_by_month(rows):
    """Clientes que aparecen por primera vez en cada mes (excluye el primer mes cargado,
    que no tiene historia previa). Devuelve {(y,m): [{name, state, salesperson}]}."""
    months = sorted(set(month_key(r['date']) for r in rows))
    if not months:
        return {}
    first_month = months[0]
    def norm(name):
        return re.sub(r"\s+", " ", str(name)).strip().lower()
    first_seen = {}
    for r in sorted(rows, key=lambda r: r['date']):
        key = norm(r['customer'])
        if key and key not in first_seen:
            first_seen[key] = r
    out = {}
    for key, r in first_seen.items():
        ym = month_key(r['date'])
        if ym == first_month:
            continue
        out.setdefault(ym, []).append({
            "name": r['customer'], "state": r['state'], "salesperson": r['salesperson'],
        })
    return out


def get_order_unit_prices(rows):
    """Precios para valuar cajas gratis, SIN pool global (estable al agregar meses):
      - order_prices[order]         : precio de caja de una línea pagada del mismo orden
      - product_month_avg[(prod,ym)]: promedio de precio de caja del producto en ese mes
      - month_avg[ym]               : promedio del mes (último recurso)
    """
    order_prices = defaultdict(float)
    pm = defaultdict(list)
    m = defaultdict(list)
    for r in rows:
        if not r['is_promo'] and r['box_price'] > 0:
            order_prices[r['order']] = r['box_price']
            ym = month_key(r['date'])
            pm[(r['product'], ym)].append(r['box_price'])
            m[ym].append(r['box_price'])
    product_month_avg = {k: sum(v) / len(v) for k, v in pm.items()}
    month_avg = {k: sum(v) / len(v) for k, v in m.items()}
    return order_prices, product_month_avg, month_avg


def promo_box_price(r, order_prices, product_month_avg, month_avg):
    """Precio por caja para valuar una caja gratis: orden → producto-mes → mes → 0."""
    if r['box_price'] > 0:
        return r['box_price']
    ym = month_key(r['date'])
    return (order_prices.get(r['order'])
            or product_month_avg.get((r['product'], ym))
            or month_avg.get(ym)
            or 0.0)


def build_monthly_state_data(rows, months, order_prices, product_avg, global_avg):
    """Build D.fl and D.ny arrays (one value per month per metric)."""
    state_month = defaultdict(lambda: defaultdict(lambda: {
        'rev': 0.0, 'cajas': 0, 'pc': 0, 'pv': 0.0, 'orders': set()
    }))

    for r in rows:
        ym = month_key(r['date'])
        if ym not in months:
            continue
        s = 'fl' if r['state'] == 'Florida' else 'ny'
        d = state_month[s][ym]
        d['cajas'] += r['qty']
        if not r['is_promo']:
            d['rev'] += r['line_rev']
            d['orders'].add(r['order'])
        else:
            # promo box: free (both amount columns are 0)
            d['pc'] += r['qty']
            d['pv'] += r['qty'] * promo_box_price(r, order_prices, product_avg, global_avg)
            d['orders'].add(r['order'])

    result = {}
    for state in ('fl', 'ny'):
        result[state] = {
            'rev': [], 'cajas': [], 'pv': [], 'pc': [], 'ord': []
        }
        for ym in months:
            d = state_month[state][ym]
            result[state]['rev'].append(round(d['rev'], 2))
            result[state]['cajas'].append(d['cajas'])
            result[state]['pv'].append(round(d['pv'], 2))
            result[state]['pc'].append(d['pc'])
            result[state]['ord'].append(len(d['orders']))
    return result


def build_period_data(rows, months, order_prices, product_avg, global_avg, month_set=None):
    """Build summary, states, customerChart, salesChart, customerTable, productsCards for a given month set."""
    if month_set is None:
        month_set = set(months)

    filtered = [r for r in rows if month_key(r['date']) in month_set]

    # --- summary & states ---
    state_data = defaultdict(lambda: {'rev': 0.0, 'cajas': 0, 'pc': 0, 'pv': 0.0, 'orders': set()})
    for r in filtered:
        s = r['state']
        d = state_data[s]
        d['cajas'] += r['qty']
        if not r['is_promo']:
            d['rev'] += r['line_rev']
            d['orders'].add(r['order'])
        else:
            # promo box: free (both amount columns are 0)
            d['pc'] += r['qty']
            d['pv'] += r['qty'] * promo_box_price(r, order_prices, product_avg, global_avg)
            d['orders'].add(r['order'])

    total_rev = sum(d['rev'] for d in state_data.values())
    total_cajas = sum(d['cajas'] for d in state_data.values())
    total_pc = sum(d['pc'] for d in state_data.values())
    total_pv = sum(d['pv'] for d in state_data.values())
    total_ord = len(set(r['order'] for r in filtered if r['order']))
    paid_cajas = total_cajas - total_pc
    avg_price = round(total_rev / paid_cajas, 2) if paid_cajas > 0 else 0

    summary = {
        'rev': round(total_rev, 2),
        'cajas': total_cajas,
        'pc': total_pc,
        'pv': round(total_pv, 2),
        'ord': total_ord,
        'avg_price': avg_price
    }

    states = {}
    for state_name, d in state_data.items():
        states[state_name] = {
            'rev': round(d['rev'], 2),
            'cajas': d['cajas'],
            'pc': d['pc'],
            'pv': round(d['pv'], 2),
            'ord': len(d['orders'])
        }
    if 'Florida' not in states:
        states['Florida'] = {'rev': 0.0, 'cajas': 0, 'pc': 0, 'pv': 0.0, 'ord': 0}
    if 'Nueva York' not in states:
        states['Nueva York'] = {'rev': 0.0, 'cajas': 0, 'pc': 0, 'pv': 0.0, 'ord': 0}

    # --- customer aggregation ---
    cust_data = defaultdict(lambda: {'rev': 0.0, 'cajas': 0, 'pc': 0, 'state': ''})
    for r in filtered:
        c = r['customer']
        cust_data[c]['state'] = r['state']
        cust_data[c]['cajas'] += r['qty']
        if not r['is_promo']:
            cust_data[c]['rev'] += r['line_rev']
        else:
            cust_data[c]['pc'] += r['qty']

    custs_sorted = sorted(cust_data.items(), key=lambda x: -x[1]['rev'])

    customer_chart = []
    for name, d in custs_sorted[:10]:
        s = 'fl' if d['state'] == 'Florida' else 'ny'
        customer_chart.append({'n': name, 'v': round(d['rev'], 2), 's': s})

    customer_table = []
    for name, d in custs_sorted:
        total_c = d['cajas']
        promo_c = d['pc']
        promo_pct = round(promo_c / total_c * 100, 1) if total_c > 0 else 0.0
        customer_table.append({
            'name': name,
            'state': d['state'],
            'revenue': round(d['rev'], 2),
            'cajas': total_c,
            'promo_qty': promo_c,
            'promo_pct': promo_pct
        })

    # --- salesperson aggregation ---
    sales_data = defaultdict(float)
    for r in filtered:
        if not r['is_promo']:
            sales_data[r['salesperson']] += r['line_rev']
    sales_sorted = sorted(sales_data.items(), key=lambda x: -x[1])
    sales_chart = [{'n': n, 'v': round(v, 2)} for n, v in sales_sorted[:10]]

    # --- products aggregation ---
    prod_data = defaultdict(lambda: {'rev': 0.0, 'cajas': 0, 'pc': 0})
    for r in filtered:
        p = r['product']
        prod_data[p]['cajas'] += r['qty']
        if not r['is_promo']:
            prod_data[p]['rev'] += r['line_rev']
        else:
            prod_data[p]['pc'] += r['qty']

    products_sorted = sorted(prod_data.items(), key=lambda x: -x[1]['rev'])
    products_cards = []
    for name, d in products_sorted:
        total_c = d['cajas']
        promo_pct = round(d['pc'] / total_c * 100, 1) if total_c > 0 else 0.0
        products_cards.append({
            'name': name,
            'revenue': round(d['rev'], 2),
            'cajas': total_c,
            'promo_pct': promo_pct
        })

    return {
        'summary': summary,
        'states': states,
        'customerChart': customer_chart,
        'salesChart': sales_chart,
        'customerTable': customer_table,
        'productsCards': products_cards
    }


def find_product_first_months(rows, months):
    """Return the first month (label) each product appears in."""
    prod_first = {}
    for ym in months:
        for r in rows:
            if month_key(r['date']) == ym:
                p = r['product']
                if p not in prod_first:
                    prod_first[p] = month_label_short(ym)
    return prod_first


def main():
    print("Leyendo fuentes/...")
    rows = load_rows()

    # Determine sorted month list
    all_months = sorted(set(month_key(r['date']) for r in rows))
    n = len(all_months)

    order_prices, product_avg, global_avg = get_order_unit_prices(rows)

    # Build D (time series)
    D = build_monthly_state_data(rows, all_months, order_prices, product_avg, global_avg)

    # Build GENERAL_DATA
    general_data = {}

    # ALL
    general_data['ALL'] = build_period_data(rows, all_months, order_prices, product_avg, global_avg,
                                            month_set=set(all_months))

    # Per month
    for ym in all_months:
        label = month_label_short(ym)
        general_data[label] = build_period_data(rows, all_months, order_prices, product_avg, global_avg,
                                                month_set={ym})

    # Find first month per product (for dynamic note in product cards)
    prod_first = find_product_first_months(rows, all_months)
    first_month_label = month_label_short(all_months[0])
    prod_first_js = {p: m for p, m in prod_first.items() if m != first_month_label}

    # MONTHS, PAIRS, PLBLS
    months_short = [month_label_short(ym) for ym in all_months]
    months_btn = [month_label_btn(ym) for ym in all_months]

    pairs = [[i, i-1] for i in range(n-1, 0, -1)]
    plbls = [[month_label_long(all_months[i]), month_label_long(all_months[i-1])]
             for i in range(n-1, 0, -1)]

    # Period pill text
    period_text = f"{month_label_btn(all_months[0])} – {month_label_btn(all_months[-1])}"

    # Serialize to JS
    js = f"""// AUTO-GENERADO por generar_data.py — no editar manualmente
// Fuente: fuentes/*.xlsx
// Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}

const MONTHS = {json.dumps(months_short, ensure_ascii=False)};
const MONTHS_BTN = {json.dumps(months_btn, ensure_ascii=False)};
const PAIRS = {json.dumps(pairs)};
const PLBLS = {json.dumps(plbls, ensure_ascii=False)};
const PERIOD_TEXT = {json.dumps(period_text, ensure_ascii=False)};
const PROD_FIRST_MONTH = {json.dumps(prod_first_js, ensure_ascii=False)};
const D = {json.dumps(D, ensure_ascii=False)};
const GENERAL_DATA = {json.dumps(general_data, ensure_ascii=False)};
"""

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(js)

    print(f"✓ {OUTPUT_FILE} generado: {period_text} ({n} meses)")

    # Quick validation
    all_summary = general_data['ALL']['summary']
    print(f"  Revenue total: ${all_summary['rev']:,.2f}")
    print(f"  Cajas: {all_summary['cajas']}")
    print(f"  % Promo: {round(all_summary['pc']/all_summary['cajas']*100,1) if all_summary['cajas'] else 0}%")


if __name__ == '__main__':
    main()
