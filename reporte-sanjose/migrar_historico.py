"""
migrar_historico.py — Migración ÚNICA del histórico de San José al formato canónico.

Lee 'Historico Ventas Latin Food - San Jose.xlsx' (esquema viejo, 13 columnas) y
escribe 'fuentes/San-Jose-historico-hasta-2026-05.xlsx' (formato canónico, 11 columnas),
para que el generar_data.py nuevo lo lea igual que cualquier otro mes.

Fiel: los montos se copian tal cual (mayo 2026 viene con Total/Unit Price intercambiados;
eso lo resuelve el parser, no este migrador). Se descarta la columna State (redundante
con Company). El Historico original queda como respaldo.

Uso: python migrar_historico.py
"""

from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
SRC = HERE / "Historico Ventas Latin Food - San Jose.xlsx"
OUT = HERE / "fuentes" / "San-Jose-historico-hasta-2026-05.xlsx"
SHEET = "San Jose"

CANON_HEADER = [
    "Order Date", "Order", "Product Variant", "Customer", "Salesperson", "Company",
    "Qty Delivered", "Qty Invoiced", "Qty Ordered", "Unit Price", "Total",
]


def migrate(src=SRC, out=OUT):
    """Reordena el Historico al formato canónico. Devuelve # de filas sin Cod Prod."""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Historico"]
    out_wb = openpyxl.Workbook()
    ows = out_wb.active
    ows.title = SHEET
    ows.append(CANON_HEADER)

    missing_cod = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        (company, state, customer, order, date, cod_prod, product,
         qd, qi, qo, salesperson, total, unit_price) = row
        if date is None:
            continue
        if not cod_prod:
            missing_cod += 1
        variant = f"{cod_prod or ''} {product or ''}".strip()
        ows.append([date, order, variant, customer, salesperson, company,
                    qd, qi, qo, unit_price, total])

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(out)
    return missing_cod


def main():
    n = migrate()
    print(f"✓ histórico migrado → {OUT.relative_to(HERE)}")
    if n:
        print(f"  ⚠ {n} filas sin Cod Prod (Product Variant quedó sin SKU)")


if __name__ == "__main__":
    main()
