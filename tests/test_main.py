import openpyxl

import split_excel as sx
from helpers import HEADER, tx, subtotal, make_workbook, D


def _seed_entrada(root):
    """Crea entrada/ con el workbook sucio bajo una ROOT temporal."""
    sheets = {
        "NY Jun - GOA": [
            HEADER,
            subtotal("Katherine (6)", qd=5),
            tx(D(2026, 6, 24), "S45227", "[GOA02] x", "Food Bazaar", "Katherine"),
            tx(D(2026, 6, 17), "S44722", "[GOA03] y", "Food Bazaar", "Katherine"),
        ],
        "MIA Jun - Jack Mckarel": [
            HEADER,
            tx(D(2026, 6, 22), "S44946", "[GET01] z", "Key Food", "Alexandra J",
               company="LatinFood Florida", qd=20, price=45.6, total=912),
            [D(2026, 5, 1), "SOLD", "FREE", "Descuentos", None, None, None, None, None, None, None],
            ["Alexandra Jimenez", 60, 10, "In Brine", 88.68, None, None, None, None, None, None],
        ],
        "NY Jun - XYZ": [
            HEADER,
            tx(D(2026, 6, 10), "S99999", "[XYZ01] mystery", "Store", "Someone"),
        ],
    }
    entrada = root / "entrada"
    entrada.mkdir()
    make_workbook(entrada / "GET Jun 2026.xlsx", sheets)


def test_main_writes_canonical_files(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(sx, "ROOT", tmp_path)
    _seed_entrada(tmp_path)

    rc = sx.main(["--no-build"])
    assert rc == 0

    goa = tmp_path / "goa-inventory" / "fuentes" / "GOA-2026-06.xlsx"
    get = tmp_path / "reporte-sanjose" / "fuentes" / "San-Jose-2026-06.xlsx"
    assert goa.exists() and get.exists()

    ws = openpyxl.load_workbook(goa)["GOA"]
    assert list(ws.iter_rows(values_only=True))[0] == tuple(sx.HEADER)
    assert ws.max_row == 3  # header + 2 líneas

    # El bloque de incentivos de la hoja de MIA debe viajar al archivo de San José
    get_ws = openpyxl.load_workbook(get).active
    cells = [c for r in get_ws.iter_rows(values_only=True) for c in r]
    assert "SOLD" in cells
    assert "Alexandra Jimenez" in cells

    out = capsys.readouterr().out
    assert "GOA" in out and "San José" in out
    assert "pendiente" in out  # GET no se buildea en SP1


def test_main_writes_unmatched(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "ROOT", tmp_path)
    _seed_entrada(tmp_path)
    sx.main(["--no-build"])
    # El XYZ sembrado es de junio → archivo namespaced por mes
    out = tmp_path / "unmatched" / "unmatched-2026-06.xlsx"
    assert out.exists()
    cells = [c for r in openpyxl.load_workbook(out).active.iter_rows(values_only=True)
             for c in r]
    assert "[XYZ01] mystery" in cells


def test_main_no_input(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(sx, "ROOT", tmp_path)
    (tmp_path / "entrada").mkdir()
    rc = sx.main(["--no-build"])
    assert rc == 1
    assert "No hay .xlsx" in capsys.readouterr().out


def test_main_only_filters_brands(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "ROOT", tmp_path)
    _seed_entrada(tmp_path)
    sx.main(["--no-build", "--only", "GOA"])
    assert (tmp_path / "goa-inventory" / "fuentes" / "GOA-2026-06.xlsx").exists()
    assert not (tmp_path / "reporte-sanjose" / "fuentes" / "San-Jose-2026-06.xlsx").exists()


def test_main_returns_1_when_generator_fails(tmp_path, monkeypatch, capsys):
    """Si el generador de una marca lista falla, main devuelve 1 (no silencia)."""
    monkeypatch.setattr(sx, "ROOT", tmp_path)
    _seed_entrada(tmp_path)
    folder = tmp_path / sx.BRANDS["GOA"].folder  # GOA es la única sp1_ready
    folder.mkdir(parents=True, exist_ok=True)
    (folder / sx.BRANDS["GOA"].generator).write_text(
        "import sys; sys.exit(2)\n", encoding="utf-8")
    rc = sx.main(["--only", "GOA"])
    assert rc == 1
    assert "generador falló" in capsys.readouterr().out
