import datetime

import openpyxl

import split_excel as sx

D = datetime.datetime


def test_write_canonical_header_rows_and_sheet(tmp_path):
    rows = [[D(2026, 6, 24), "S45227", "[GOA02] x", "c", "sp",
             "LatinFood US Corp.", 1, 1, 1, 21.85, 21.85]]
    out = tmp_path / "GOA-2026-06.xlsx"
    sx.write_canonical(out, rows, [], "GOA")

    wb = openpyxl.load_workbook(out)
    ws = wb["GOA"]
    grid = list(ws.iter_rows(values_only=True))
    assert list(grid[0]) == sx.HEADER
    assert grid[1][1] == "S45227"
    assert isinstance(grid[1][0], datetime.datetime)


def test_write_canonical_preserves_fractional_qty(tmp_path):
    rows = [[D(2026, 6, 18), "S44804", "[KOM01] x", "c", "sp",
             "LatinFood US Corp.", 0.5, 0.5, 0.5, 50.64, 25.32]]
    out = tmp_path / "KOM-2026-06.xlsx"
    sx.write_canonical(out, rows, [], "KOM")

    ws = openpyxl.load_workbook(out).active
    assert list(ws.iter_rows(values_only=True))[1][6] == 0.5


def test_write_canonical_appends_incentive_block(tmp_path):
    rows = [[D(2026, 6, 22), "S44946", "[GET01] x", "c", "sp",
             "LatinFood Florida", 20, 20, 20, 45.6, 912]]
    block = [["Alexandra Jimenez", 60, 10, "In Brine == > 3 cases", 88.68,
              None, None, None, None, None, None]]
    out = tmp_path / "San-Jose-2026-06.xlsx"
    sx.write_canonical(out, rows, block, "San Jose")

    grid = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # header + 1 fila + separador en blanco + 1 fila de bloque
    assert len(grid) == 4
    assert all(c is None for c in grid[2])            # separador
    assert grid[3][0] == "Alexandra Jimenez"


def test_write_canonical_creates_missing_dirs(tmp_path):
    out = tmp_path / "nueva" / "fuentes" / "GOA-2026-06.xlsx"
    sx.write_canonical(out, [], [], "GOA")
    assert out.exists()
