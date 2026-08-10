"""
report_core.py — Núcleo compartido para generar el `data.js` de una marca a partir
de su carpeta `fuentes/` (formato canónico de 11 columnas del distribuidor).

Es la generalización del generador de GOA (`goa-inventory/parse_excel.py`):
mismo objeto `reportData`, pero parametrizado por marca / prefijo de SKU / hoja, y
**preservando cantidades fraccionarias** (algunas marcas venden medias cajas: 0.5).

Todo cálculo derivado vive aquí en funciones explícitas; el dashboard solo formatea.

Uso desde el wrapper de una marca:

    import report_core
    report_core.build_and_write(
        here=Path(__file__).parent,
        brand="Kombuchacha", distributor="LatinFood US Corp",
        sku_prefix="KOM", sheet_candidates=("KOM", "Sheet1"))
"""

import calendar
import re
import json
import datetime
from pathlib import Path

import openpyxl

from column_resolver import (CANONICAL_POSITIONS, find_header_row, get,
                             resolve_columns)

# Índices de columnas (0-based) en el bloque transaccional canónico
COL_DATE, COL_ORDER, COL_VARIANT, COL_CUSTOMER, COL_SALESPERSON = 0, 1, 2, 3, 4
COL_COMPANY, COL_QTY_DELIVERED, COL_QTY_INVOICED, COL_QTY_ORDERED = 5, 6, 7, 8
COL_UNIT_PRICE, COL_TOTAL = 9, 10

MESES_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

# Mes de un canónico mensual por nombre de archivo ('KOM-2026-07.xlsx' -> '2026-07').
# Un archivo así SIN líneas es un mes en cero explícito (D3: "vendió $0" y "no
# cargué el mes" deben verse distinto); los crudos del distribuidor no matchean.
FILENAME_MONTH_RE = re.compile(r"-(\d{4}-\d{2})\.xlsx$", re.IGNORECASE)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de limpieza / parametrización
# ─────────────────────────────────────────────────────────────────────────────
def num(value):
    """Convierte a número (float) de forma segura. None/texto/bool/NaN -> 0."""
    if value is None or isinstance(value, bool):
        return 0
    if isinstance(value, (int, float)):
        return 0 if value != value else value  # descarta NaN
    return 0


def q(value):
    """Cantidad: número redondeado a 2 decimales (preserva 0.5, evita ruido float)."""
    return round(num(value), 2)


def clean_str(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sku_pattern(prefix):
    """Patrón para '[PREFIJO<n>] Nombre' -> ('PREFIJO<n>', 'Nombre')."""
    return re.compile(rf"^\s*\[({re.escape(prefix)}\d+)\]\s*(.*)$")


def split_sku(variant, pattern):
    m = pattern.match(str(variant))
    if not m:
        return None, clean_str(variant)
    return m.group(1), clean_str(m.group(2))


def month_key(date_iso):
    return date_iso[:7]


def month_label(key):
    y, m = key.split("-")
    return f"{MESES_ES[int(m)].capitalize()} {y}"


def list_source_files(folder):
    folder = Path(folder)
    if not folder.exists():
        return []
    return sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~$"))


def read_sheet_rows(path, sheet_candidates):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in sheet_candidates:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    return list(ws.iter_rows(values_only=True))


def is_transaction_row(row, pattern, cols):
    date = get(row, cols, "date")
    variant = get(row, cols, "variant")
    return (
        isinstance(date, datetime.datetime)
        and isinstance(variant, str)
        and pattern.match(variant)
    )


def parse_transactions(rows, pattern, cols=None):
    """Devuelve la lista de líneas de pedido limpias (dicts).

    `cols` mapea campo lógico -> índice (ver column_resolver). Si es None, se
    resuelve desde el encabezado de `rows`; sin encabezado, posiciones canónicas.
    """
    if cols is None:
        hdr_i = find_header_row(rows)
        cols = (resolve_columns(rows[hdr_i]) if hdr_i is not None
                else dict(CANONICAL_POSITIONS))
    lines = []
    for row in rows:
        if not is_transaction_row(row, pattern, cols):
            continue
        sku, name = split_sku(get(row, cols, "variant"), pattern)
        if sku is None:
            continue
        date = get(row, cols, "date")
        lines.append({
            "date": date.date().isoformat(),
            "month": month_key(date.date().isoformat()),
            "order": clean_str(get(row, cols, "order")),
            "sku": sku,
            "product": name,
            "customer": clean_str(get(row, cols, "customer")),
            "salesperson": clean_str(get(row, cols, "salesperson")),
            "qtyOrdered": q(get(row, cols, "qty_ordered")),
            "qtyDelivered": q(get(row, cols, "qty_delivered")),
            "qtyInvoiced": q(get(row, cols, "qty_invoiced")),
            "unitPrice": round(num(get(row, cols, "unit_price")), 2),
            "total": round(num(get(row, cols, "total")), 2),
        })
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Incentivos (bloque manual del final; tolerante a fallos) — igual que GOA
# ─────────────────────────────────────────────────────────────────────────────
def empty_incentives():
    return {"freeUnitsBySalesperson": [], "costItems": [], "totalDescuentos": 0.0}


def merge_incentives(incs):
    free, cost, total = {}, [], 0.0
    for inc in incs:
        for f in inc.get("freeUnitsBySalesperson", []):
            free[f["name"]] = free.get(f["name"], 0) + f["free"]
        cost.extend(inc.get("costItems", []))
        total += inc.get("totalDescuentos", 0.0)
    return {
        "freeUnitsBySalesperson": [{"name": n, "free": v} for n, v in free.items()],
        "costItems": cost,
        "totalDescuentos": round(total, 2),
    }


def dominant_month(lines):
    counts = {}
    for ln in lines:
        counts[ln["month"]] = counts.get(ln["month"], 0) + 1
    return max(counts, key=counts.get)


def _last_number(row):
    for cell in reversed(list(row)):
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            if cell == cell:  # no NaN
                return float(cell)
    return None


def _looks_like_cost_label(text):
    t = text.upper()
    return any(k in t for k in ("CAJA", "MUESTRA", "NEW CUSTOMER", "DESCUENTO", "INCENTIVO"))


def _find_cost_label(row):
    for cell in row:
        s = clean_str(cell)
        if s and _looks_like_cost_label(s):
            return s
    return None


def parse_incentives(rows):
    free_by_sp = []
    cost_items = []
    total_descuentos = 0.0

    start = None
    for i, row in enumerate(rows):
        if clean_str(row[COL_ORDER]).upper() == "SOLD":
            start = i
            break
    if start is None:
        return empty_incentives()

    header = rows[start]
    free_col = COL_VARIANT  # por defecto C
    for c, cell in enumerate(header):
        if clean_str(cell).upper() == "FREE":
            free_col = c
            break

    for row in rows[start + 1:]:
        label = clean_str(row[COL_DATE]) if len(row) else ""
        last_num = _last_number(row)

        free_val = num(row[free_col]) if len(row) > free_col else 0
        is_sp_row = bool(label) and not _looks_like_cost_label(label)
        if is_sp_row and free_val > 0:
            free_by_sp.append({"name": label, "free": int(free_val)})

        cost_label = _find_cost_label(row)
        if cost_label and last_num:
            cost_items.append({"label": cost_label, "amount": round(last_num, 2)})

        if last_num:
            total_descuentos = round(last_num, 2)

    return {
        "freeUnitsBySalesperson": free_by_sp,
        "costItems": cost_items,
        "totalDescuentos": total_descuentos,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Lectura de fuentes
# ─────────────────────────────────────────────────────────────────────────────
def load_all_sources(folder, pattern, sheet_candidates):
    """Lee todas las fuentes, dedup por orden entre archivos, e incentivos por mes.

    Devuelve (lines, incentives_by_month, source_names, dup_orders, named_months):
    `named_months` son los meses declarados por nombre de archivo canónico mensual.
    """
    seen_orders = set()
    lines = []
    incentives_by_month = {}
    dup_orders = []
    sources = []
    named_months = set()

    for path in list_source_files(folder):
        try:
            rows = read_sheet_rows(path, sheet_candidates)
        except Exception as e:
            print(f"  ⚠ No se pudo leer {path.name}: {e} (se omite)")
            continue
        sources.append(path.name)
        m = FILENAME_MONTH_RE.search(path.name)
        if m:
            named_months.add(m.group(1))
        hdr_i = find_header_row(rows)
        cols = (resolve_columns(rows[hdr_i], label=path.name)
                if hdr_i is not None else dict(CANONICAL_POSITIONS))
        file_lines = parse_transactions(rows, pattern, cols)
        kept = []
        for ln in file_lines:
            if ln["order"] in seen_orders:
                dup_orders.append(ln["order"])
                continue
            kept.append(ln)
        for ln in kept:
            seen_orders.add(ln["order"])
        lines.extend(kept)

        inc = parse_incentives(rows)
        has_inc = inc["freeUnitsBySalesperson"] or inc["costItems"] or inc["totalDescuentos"]
        if has_inc and kept:
            m = dominant_month(kept)
            incentives_by_month[m] = merge_incentives(
                [incentives_by_month.get(m, empty_incentives()), inc])

    return lines, incentives_by_month, sources, sorted(set(dup_orders)), named_months


# ─────────────────────────────────────────────────────────────────────────────
# Agregaciones (unidades como float, redondeadas a 2 decimales)
# ─────────────────────────────────────────────────────────────────────────────
def aggregate_products(lines):
    prods = {}
    for ln in lines:
        p = prods.setdefault(ln["sku"], {
            "sku": ln["sku"], "name": ln["product"],
            "units": 0.0, "revenue": 0.0, "_orders": set(), "_sp": {},
        })
        p["units"] += ln["qtyOrdered"]
        p["revenue"] += ln["total"]
        p["_orders"].add(ln["order"])
        sp = p["_sp"].setdefault(ln["salesperson"], {"units": 0.0, "revenue": 0.0})
        sp["units"] += ln["qtyOrdered"]
        sp["revenue"] += ln["total"]

    out = []
    for p in prods.values():
        salespeople = [
            {"name": n, "units": round(v["units"], 2), "revenue": round(v["revenue"], 2)}
            for n, v in p["_sp"].items()
        ]
        salespeople.sort(key=lambda s: (-s["units"], -s["revenue"]))
        out.append({
            "sku": p["sku"], "name": p["name"],
            "units": round(p["units"], 2), "revenue": round(p["revenue"], 2),
            "orders": len(p["_orders"]),
            "salespeople": salespeople,
        })
    out.sort(key=lambda p: (-p["units"], -p["revenue"]))
    return out


def aggregate_salespeople(lines):
    sps = {}
    for ln in lines:
        s = sps.setdefault(ln["salesperson"], {
            "name": ln["salesperson"], "units": 0.0, "revenue": 0.0,
            "_orders": set(), "_customers": set(),
        })
        s["units"] += ln["qtyOrdered"]
        s["revenue"] += ln["total"]
        s["_orders"].add(ln["order"])
        s["_customers"].add(ln["customer"])

    out = [{
        "name": s["name"], "units": round(s["units"], 2), "revenue": round(s["revenue"], 2),
        "orders": len(s["_orders"]), "customers": len(s["_customers"]),
    } for s in sps.values()]
    out.sort(key=lambda s: (-s["revenue"], -s["units"]))
    return out


def aggregate_customers(lines):
    cs = {}
    for ln in lines:
        c = cs.setdefault(ln["customer"], {
            "name": ln["customer"], "units": 0.0, "revenue": 0.0, "_orders": set(),
        })
        c["units"] += ln["qtyOrdered"]
        c["revenue"] += ln["total"]
        c["_orders"].add(ln["order"])

    out = [{
        "name": c["name"], "units": round(c["units"], 2), "revenue": round(c["revenue"], 2),
        "orders": len(c["_orders"]),
    } for c in cs.values()]
    out.sort(key=lambda c: (-c["revenue"], -c["units"]))
    return out


def aggregate_daily(lines):
    days = {}
    for ln in lines:
        d = days.setdefault(ln["date"], {
            "date": ln["date"], "units": 0.0, "revenue": 0.0, "_orders": set(),
        })
        d["units"] += ln["qtyOrdered"]
        d["revenue"] += ln["total"]
        d["_orders"].add(ln["order"])

    out = [{
        "date": d["date"], "units": round(d["units"], 2),
        "revenue": round(d["revenue"], 2), "orders": len(d["_orders"]),
    } for d in days.values()]
    out.sort(key=lambda d: d["date"])
    return out


def build_view(lines, incentives):
    products = aggregate_products(lines)
    salespeople = aggregate_salespeople(lines)
    customers = aggregate_customers(lines)
    daily = aggregate_daily(lines)

    free_map = {f["name"].lower(): f["free"] for f in incentives["freeUnitsBySalesperson"]}
    for s in salespeople:
        s["freeUnits"] = next(
            (v for k, v in free_map.items()
             if k in s["name"].lower() or s["name"].lower().startswith(k[:8])),
            0,
        )

    totals = {
        "units": round(sum(p["units"] for p in products), 2),
        "revenue": round(sum(p["revenue"] for p in products), 2),
        "orders": len({ln["order"] for ln in lines}),
        "customers": len({ln["customer"] for ln in lines}),
        "salespeople": len(salespeople),
        "freeUnits": sum(f["free"] for f in incentives["freeUnitsBySalesperson"]),
    }
    return {
        "totals": totals, "products": products, "salespeople": salespeople,
        "customers": customers, "daily": daily, "incentives": incentives,
    }


def build_months(lines, incentives_by_month, empty_month_keys=()):
    by_month = {}
    for ln in lines:
        by_month.setdefault(ln["month"], []).append(ln)

    out = []
    for key in sorted(by_month):
        month_lines = by_month[key]
        inc = incentives_by_month.get(key, empty_incentives())
        view = build_view(month_lines, inc)
        dates = sorted(ln["date"] for ln in month_lines)
        out.append({
            "key": key,
            "label": month_label(key),
            "periodStart": dates[0],
            "periodEnd": dates[-1],
            **view,
        })

    # Meses en cero: declarados por archivo canónico mensual pero sin líneas.
    for key in sorted(set(empty_month_keys) - set(by_month)):
        y, mo = map(int, key.split("-"))
        last_day = calendar.monthrange(y, mo)[1]
        out.append({
            "key": key,
            "label": month_label(key),
            "periodStart": f"{key}-01",
            "periodEnd": f"{key}-{last_day:02d}",
            "empty": True,
            **build_view([], empty_incentives()),
        })

    out.sort(key=lambda m: m["key"])
    return out


def merge_views(views):
    prods = {}
    for v in views:
        for p in v["products"]:
            t = prods.setdefault(p["sku"], {
                "sku": p["sku"], "name": p["name"],
                "units": 0.0, "revenue": 0.0, "orders": 0, "_sp": {}})
            t["units"] += p["units"]; t["revenue"] += p["revenue"]; t["orders"] += p["orders"]
            for sp in p["salespeople"]:
                s = t["_sp"].setdefault(sp["name"], {"units": 0.0, "revenue": 0.0})
                s["units"] += sp["units"]; s["revenue"] += sp["revenue"]
    products = []
    for t in prods.values():
        sps = [{"name": n, "units": round(s["units"], 2), "revenue": round(s["revenue"], 2)}
               for n, s in t["_sp"].items()]
        sps.sort(key=lambda s: (-s["units"], -s["revenue"]))
        products.append({"sku": t["sku"], "name": t["name"], "units": round(t["units"], 2),
                         "revenue": round(t["revenue"], 2), "orders": t["orders"],
                         "salespeople": sps})
    products.sort(key=lambda p: (-p["units"], -p["revenue"]))

    sps = {}
    for v in views:
        for s in v["salespeople"]:
            t = sps.setdefault(s["name"], {"name": s["name"], "units": 0.0, "revenue": 0.0,
                                           "orders": 0, "customers": 0, "freeUnits": 0})
            t["units"] += s["units"]; t["revenue"] += s["revenue"]; t["orders"] += s["orders"]
            t["customers"] += s["customers"]; t["freeUnits"] += s.get("freeUnits", 0)
    salespeople = [{**t, "units": round(t["units"], 2), "revenue": round(t["revenue"], 2)}
                   for t in sps.values()]
    salespeople.sort(key=lambda s: (-s["revenue"], -s["units"]))

    cs = {}
    for v in views:
        for c in v["customers"]:
            t = cs.setdefault(c["name"], {"name": c["name"], "units": 0.0, "revenue": 0.0, "orders": 0})
            t["units"] += c["units"]; t["revenue"] += c["revenue"]; t["orders"] += c["orders"]
    customers = [{**t, "units": round(t["units"], 2), "revenue": round(t["revenue"], 2)}
                 for t in cs.values()]
    customers.sort(key=lambda c: (-c["revenue"], -c["units"]))

    daily = sorted((d for v in views for d in v["daily"]), key=lambda d: d["date"])
    incentives = merge_incentives([v["incentives"] for v in views])

    totals = {
        "units": round(sum(p["units"] for p in products), 2),
        "revenue": round(sum(p["revenue"] for p in products), 2),
        "orders": sum(v["totals"]["orders"] for v in views),
        "customers": len(cs),
        "salespeople": len(sps),
        "freeUnits": sum(f["free"] for f in incentives["freeUnitsBySalesperson"]),
    }
    return {"totals": totals, "products": products, "salespeople": salespeople,
            "customers": customers, "daily": daily, "incentives": incentives}


def load_previous_report(path):
    path = Path(path)
    if not path.exists():
        return None
    txt = path.read_text(encoding="utf-8")
    m = re.search(r"const reportData = (.*);\s*$", txt, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None


def carry_forward(months, prev_report):
    if not prev_report or "months" not in prev_report:
        return months, []
    present = {m["key"] for m in months}
    carried = []
    for pm in prev_report["months"]:
        if pm["key"] not in present:
            months.append(pm)
            carried.append(pm["key"])
    months.sort(key=lambda m: m["key"])
    return months, sorted(carried)


# ─────────────────────────────────────────────────────────────────────────────
# Orquestación
# ─────────────────────────────────────────────────────────────────────────────
def build_report(here, brand, distributor, sku_prefix, sheet_candidates):
    here = Path(here)
    sources_dir = here / "fuentes"
    output_file = here / "data.js"
    pattern = sku_pattern(sku_prefix)

    lines, inc_by_month, sources, dup_orders, named_months = load_all_sources(
        sources_dir, pattern, sheet_candidates)
    if not lines:
        raise RuntimeError(f"No se detectaron líneas en {sources_dir}. ¿Pusiste los Excel?")

    months = build_months(lines, inc_by_month, empty_month_keys=named_months)
    prev = load_previous_report(output_file)
    months, carried = carry_forward(months, prev)

    general = build_view(lines, merge_incentives(list(inc_by_month.values())))
    if carried:
        carried_views = [
            {k: m[k] for k in ("totals", "products", "salespeople",
                               "customers", "daily", "incentives")}
            for m in months if m["key"] in carried
        ]
        general = merge_views([general] + carried_views)

    period_start = min(m["periodStart"] for m in months)
    period_end = max(m["periodEnd"] for m in months)
    period_label = (months[0]["label"] if len(months) == 1
                    else f"{months[0]['label']} – {months[-1]['label']}")

    return {
        "meta": {
            "brand": brand,
            "distributor": distributor,
            "periodStart": period_start,
            "periodEnd": period_end,
            "periodLabel": period_label,
            "lineItems": len(lines),
            "sources": sources,
            "carriedForward": carried,
            "duplicateOrders": dup_orders,
        },
        **general,
        "months": months,
    }


def build_and_write(here, brand, distributor, sku_prefix, sheet_candidates):
    report = build_report(here, brand, distributor, sku_prefix, sheet_candidates)
    output_file = Path(here) / "data.js"
    js = "const reportData = " + json.dumps(report, indent=2, ensure_ascii=False) + ";\n"
    output_file.write_text(js, encoding="utf-8")

    m, t = report["meta"], report["totals"]
    print("✓ data.js generado")
    print(f"  Marca:      {m['brand']}")
    print(f"  Periodo:    {m['periodLabel']} ({m['periodStart']} → {m['periodEnd']})")
    print(f"  Fuentes:    {', '.join(m['sources']) or '(ninguna)'}")
    print(f"  Meses:      {', '.join(mo['key'] for mo in report['months'])}")
    print(f"  Líneas:     {m['lineItems']}")
    print(f"  Ingresos:   ${t['revenue']:,.2f}   Unidades: {t['units']}   Órdenes: {t['orders']}")
    if m["duplicateOrders"]:
        print(f"  ⚠ Órdenes duplicadas omitidas: {', '.join(m['duplicateOrders'])}")
    if m["carriedForward"]:
        print(f"  ⚠ Meses conservados: {', '.join(m['carriedForward'])}")
    return report
